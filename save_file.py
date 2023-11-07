import pandas as pd
import constants as const
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from docx.shared import Cm
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Cm, Inches, Mm, Emu

def write_excel(incr: dict, percent: dict, eater: dict, top: dict):
    print("\nPopulating G-Sheet Template...")
    wb = load_workbook(const.g_sheet_template)

    incr_sheet = wb[const.camp_incr_sheet]
    # populating data to respective cells
    populate_cells(incr_sheet, incr[const.dict_pre_vs_test_table], 27, 1, True)
    populate_cells(incr_sheet, incr[const.dict_camp_incr_data], 5, 1, True)
    attach_graph(incr_sheet, const.camp_incr_graph, "I5", 650, 800)

    top_perf_sheet = wb[const.top_perf_sheet]
    populate_cells(top_perf_sheet, top[const.dict_top_cities_data], 45, 3, True)
    populate_cells(top_perf_sheet, top[const.dict_top_restos_data], 45, 13, True)
    attach_graph(top_perf_sheet, const.top_cities_graph, "A4", 650, 800)
    attach_graph(top_perf_sheet, const.top_restos_graph, "L4", 750, 1000)

    eater_sheet = wb[const.eaters_sheet]
    uber_one_sheet = wb[const.uber_one_sheet]
    populate_cells(eater_sheet, eater[const.dict_eaters_data], 2, 1, True)
    populate_cells(eater_sheet, eater[const.dict_eaters_table], 3, 19, True)
    populate_cells(uber_one_sheet, eater[const.dict_uber_one_data], 2, 1, True)
    attach_graph(eater_sheet, const.order_new_eater_graph, "R8", 500, 700)
    attach_graph(eater_sheet, const.abs_new_eater_graph, "R35", 500, 700)
    attach_graph(eater_sheet, const.sales_new_eater_graph, "R65", 500, 700)

    percent_sheet = wb[const.percent_sheet]
    summary_sheet = wb[const.summary_sheet]
    ads_manager_sheet = wb[const.ads_manager_sheet]
    ads_metrics_sheet = wb[const.ads_metrics_sheet]
    overall_sheet = wb[const.overall_perf_sheet]
    populate_cells(summary_sheet, percent[const.summary_sheet], 4, 3, False)
    populate_cells(ads_manager_sheet, percent[const.dict_ads_manager_data], 1, 1, True)
    populate_cells(ads_metrics_sheet, percent[const.dict_ads_metrics_data], 2, 1, True)
    populate_cells(overall_sheet, percent[const.dict_overall_perf_data], 2, 1, True)
    populate_cells(percent_sheet, percent["Impression_data"], 3, 1, False)
    populate_cells(percent_sheet, percent["Impression"], 8, 1, True)
    populate_cells(percent_sheet, percent["Clicks_data"], 3, 5, False)
    populate_cells(percent_sheet, percent["Clicks"], 8, 5, True)   
    populate_cells(percent_sheet, percent["Orders_data"], 3, 9, False)
    populate_cells(percent_sheet, percent["Orders"], 8, 9, True)
    populate_cells(percent_sheet, percent["Sales_data"], 3, 13, False)
    populate_cells(percent_sheet, percent["Sales"], 8, 13, True)
    populate_cells(percent_sheet, percent["Conversion Rate"], 28, 2, True)
    
    attach_graph(percent_sheet, const.impressions_graph, "A12", 300, 375)
    attach_graph(percent_sheet, const.clicks_graph, "E12", 300, 375)
    attach_graph(percent_sheet, const.orders_graph, "I12", 300, 375)
    attach_graph(percent_sheet, const.sales_graph, "M12", 300, 375)
    attach_graph(percent_sheet, const.conversion_graph, "A32", 300, 375)
    #saving excel
    wb.save(const.gsheet)
    # wb.save("/Users/rramak2/Documents/projects/pcr/output/gsheet.xlsx")

def populate_cells(ws, table, start_row, start_col, header):
    rows = dataframe_to_rows(table, index=False, header=header)
    for r_idx, row in enumerate(rows, start_row):
            for c_idx, value in enumerate(row, start_col):
                ws.cell(row=r_idx, column=c_idx, value=value)

def attach_graph(ws, graph, anchor, height, width):
     # fetching the graph from local storage
    img = openpyxl.drawing.image.Image(graph)
    img.height = height
    img.width = width
    # attaching the image to a cell
    img.anchor = anchor
    #adding the image to sheet
    ws.add_image(img)
 


def write_doc(incr: dict, percent: dict, eater: dict, top: dict):
    template = DocxTemplate(const.tldr_template)
    content = dict()

    content["starttest"] = incr["teststart"][:11]
    content["endtest"] = incr["testend"][:11]
    content["campname"] = percent[const.dict_campaign_name]
    content["revenue"] = percent[const.summary_sheet].iat[0,1]
    content["roas"] = percent[const.summary_sheet].iat[1,1]
    content["orders"] = percent[const.summary_sheet].iat[2,1]
    content["cpr"] = percent[const.summary_sheet].iat[3,1]
    content["avb"] = percent[const.summary_sheet].iat[4,1]
    content["cvr"] = percent[const.summary_sheet].iat[5,1]
    content["impressions"] = percent[const.summary_sheet].iat[6,1]
    content["clicks"] = percent[const.summary_sheet].iat[7,1]
    content["ctr"] = percent[const.summary_sheet].iat[8,1]
    content["cpc"] = percent[const.summary_sheet].iat[9,1]
    content["orgimp"] = percent["Impression"].iat[0,1]
    content["adimp"] = percent["Impression"].iat[0,2]
    content["orgimpp"] = percent["Impression"].iat[1,1]
    content["adimpp"] = percent["Impression"].iat[1,2]
    populate_image_content(template, content, "grimp", const.impressions_graph, 4, 5.5)
    content["orgclick"] = percent["Clicks"].iat[0,1]
    content["adclick"] = percent["Clicks"].iat[0,2]
    content["orgclickp"] = percent["Clicks"].iat[1,1]
    content["adclickp"] = percent["Clicks"].iat[1,2]
    populate_image_content(template, content, "grclick", const.clicks_graph, 4, 5.5)
    content["orgorder"] = percent["Orders"].iat[0,1]
    content["adorder"] = percent["Orders"].iat[0,2]
    content["orgorderp"] = percent["Orders"].iat[1,1]
    content["adorderp"] = percent["Orders"].iat[1,2]
    populate_image_content(template, content, "grorder", const.orders_graph, 4, 5.5)
    content["orgsales"] = percent["Sales"].iat[0,1]
    content["adsales"] = percent["Sales"].iat[0,2]
    content["orgsalesp"] = percent["Sales"].iat[1,1]
    content["adsalesp"] = percent["Sales"].iat[1,2]
    populate_image_content(template, content, "grsales", const.sales_graph, 4, 5.5)   
    content["orgconvp"] = percent["Conversion Rate"].iat[0,0]
    content["adconvp"] = percent["Conversion Rate"].iat[0,1]
    populate_image_content(template, content, "grconv", const.impressions_graph, 4, 5.5)   
    
 




    template.render(content)
    template.save(const.tldr)

def populate_image_content(template, content: dict, variable, img, height, width):
    #attaching the image to template with image orientations
    image = InlineImage(template, img,Cm(10))
    image.height = Inches(height)
# 6.5   4.01
# top store     7.6 .    3.54 4.69
    image.width = Inches(width)
    #assigning image to it's variable for doc
    content[variable] = image
    return content

